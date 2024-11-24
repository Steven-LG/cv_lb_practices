import cv2
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import os

# ! Cambiar valores
RANGO = 30
# IMAGE_FILE = os.path.join('img','/home/xsvd/Main/Homework/LBVC/fifth_practice/image.png')


def encontrar_vecinos(imagen, fila, columna,etiquetas, etiqueta_actual, pixel_origen):
    vecinos = []

    stack = [(fila, columna)]

    while stack:
        fila, columna = stack.pop()
        for i in range(-1, 2):
            for j in range(-1, 2):
                if(i == 0 and j== 0):
                    continue

                vecino_fila = fila + i
                vecino_columna = columna + j

                if (0 <= vecino_fila < imagen.shape[0]) and (0 <= vecino_columna < imagen.shape[1]) and (i != 0 or j != 0):

                    pixel_vecino = imagen[vecino_fila, vecino_columna]
                    if abs(int(pixel_vecino - pixel_origen)) <= RANGO:
                        if (vecino_fila, vecino_columna) not in etiquetas:
                            etiquetas[(vecino_fila, vecino_columna)] = etiqueta_actual
                            stack.append((vecino_fila, vecino_columna))
                        elif etiquetas[(vecino_fila, vecino_columna)] != etiqueta_actual:
                            vecinos.append((vecino_fila, vecino_columna))
    return vecinos

def etiquetar_grupos(imagen):
    etiquetas = {}
    etiqueta_actual = 1

    for fila in range(imagen.shape[0]):
        for columna in  range(imagen.shape[1]):
            if (fila, columna) not in etiquetas:
                pixel_origen = int(imagen[fila,columna])
                vecinos = encontrar_vecinos(imagen, fila, columna, etiquetas, etiqueta_actual, pixel_origen)
                if vecinos:
                    etiquetas[(fila, columna)] = etiqueta_actual
                    for vecino in vecinos:
                        if vecino not in etiquetas:
                            etiquetas[vecino] = etiqueta_actual
                # else:
                #     etiquetas[(fila, columna)] = 0
                etiqueta_actual += 1
    return etiquetas


def convertir_a_excel(df):

    df = pd.DataFrame(array)
    wb = Workbook()

    # Seleccionar la primera hoja del libro de trabajo
    ws = wb.active

    for index, row in df.iterrows():

        ws.cell(row['Fila']+1, row['Columna']+1, row['Valor'])

    # ! Codigo para colorear celdas, no se pudo completar esta parte ;-;
    wb.save('./output/vecindarios.xlsx')


# Cargar imagen en escala de grises
imagen_gris = cv2.imread("/home/xsvd/Main/Homework/LBVC/fifth_practice/image.png", cv2.IMREAD_GRAYSCALE)
etiquetas = etiquetar_grupos(imagen_gris)
print("Número de grupos encontrados:", max(etiquetas.values()))

array = []
for coordenada, valor in etiquetas.items():
    fila, columna = coordenada
    # array[fila, columna] = valor
    array.append({'Fila': fila, 'Columna': columna, 'Valor': valor})

convertir_a_excel(array)